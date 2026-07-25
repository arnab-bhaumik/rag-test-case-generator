"""Writes approved test cases (+ the RTM) to a formatted .xlsx workbook."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.schemas import Run, ReviewStatus
from src.traceability.rtm_builder import build_rtm

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TC_COLUMNS = [
    ("Test Case ID", 16),
    ("Test Scenario", 34),
    ("Description", 34),
    ("Category", 13),
    ("Priority", 10),
    ("Module", 12),
    ("Pre-conditions", 30),
    ("Steps", 44),
    ("Expected Results", 34),
    ("Trace", 18),
    ("Status", 11),
    ("Jira Key", 12),
]

RTM_COLUMNS = [("AC", 12), ("Requirement / Condition", 44), ("Linked Test Cases", 30), ("Covered", 12)]


def _write_header(ws: Worksheet, columns: list[tuple[str, int]]) -> None:
    for i, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def build_workbook(run: Run) -> Workbook:
    approved = [tc for tc in run.test_cases if tc.status == ReviewStatus.approved]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    _write_header(ws, TC_COLUMNS)

    for row, tc in enumerate(approved, start=2):
        values = [
            tc.id,
            tc.title,
            tc.description,
            tc.category.value if tc.category else "",
            tc.priority.value,
            tc.module or "",
            tc.preconditions,
            "\n".join(f"{i + 1}. {s}" for i, s in enumerate(tc.steps)),
            tc.expected_result,
            tc.trace or "",
            tc.status.value,
            tc.jira_key or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("RTM")
    _write_header(ws2, RTM_COLUMNS)
    for row, r in enumerate(build_rtm(run.conditions, run.test_cases), start=2):
        values = [r.ac_ref or "—", r.condition_text, ", ".join(r.linked_test_case_ids), "Covered" if r.covered else "Not covered"]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def export_to_bytes(run: Run) -> bytes:
    buf = io.BytesIO()
    build_workbook(run).save(buf)
    return buf.getvalue()
