"""Parses old test cases from CSV or Excel into the TestCase schema."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from src.models.schemas import TestCase

# Each schema field accepts any of these header spellings — matched case/whitespace/
# hyphen-insensitively (see _normalize()), first alias found in the file wins. This
# tolerance matters: real-world exports vary ("Pre-conditions" vs "Preconditions",
# "Test Steps" vs "Steps"), and an exact-string mismatch used to silently drop that
# column's data entirely rather than fail loudly.
FIELD_ALIASES: dict[str, list[str]] = {
    "id": ["Test Case ID", "ID", "Test ID", "TC ID"],
    "title": ["Test Scenario", "Title", "Test Case Title", "Scenario", "Summary"],
    "description": ["Description", "Test Case Description", "Test Objective", "Objective"],
    "preconditions": ["Pre-conditions", "Preconditions", "Precondition", "Pre-requisites", "Prerequisites"],
    "steps": ["Steps", "Test Steps", "Step", "Test Step"],
    "expected_result": ["Expected Results", "Expected Result", "Expected Outcome", "Expected"],
    "module": ["Module", "Component", "Feature"],
    "priority": ["Priority"],
}

_STEP_PREFIX_RE = re.compile(r"^\s*(?:\d+[.)]|[-*•])\s*")


def _normalize(header: str) -> str:
    return re.sub(r"[\s\-_]+", " ", header.strip().lower())


def _resolve_column_map(headers: list[str]) -> dict[str, str]:
    """Matches each schema field to whichever actual column header is present in
    this file. Returns {field: actual_header_text}; a field with no matching
    column is simply absent (get() in _row_to_test_case treats it as blank)."""
    normalized_to_actual = {_normalize(h): h for h in headers if h}
    resolved: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            actual = normalized_to_actual.get(_normalize(alias))
            if actual:
                resolved[field] = actual
                break
    return resolved


def _split_steps(raw: str) -> list[str]:
    """Splits a multi-line Steps cell into individual steps, stripping any manual
    numbering/bullets the source file already had ("1. ", "2)", "- ", "• ") so the
    UI's own numbering doesn't double up on top of it."""
    if not raw:
        return []
    lines = [_STEP_PREFIX_RE.sub("", line).strip() for line in raw.splitlines() if line.strip()]
    return lines or [raw.strip()]


def _row_to_test_case(row: dict[str, str], column_map: dict[str, str]) -> TestCase:
    def get(key: str) -> str:
        col = column_map.get(key)
        return (row.get(col) or "").strip() if col else ""

    return TestCase(
        id=get("id"),
        title=get("title"),
        description=get("description"),
        preconditions=get("preconditions"),
        steps=_split_steps(get("steps")),
        expected_result=get("expected_result"),
        module=get("module") or None,
        priority=get("priority") or "Medium",
        source="library",
    )


def load_csv(path: str | Path) -> list[TestCase]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        column_map = _resolve_column_map(reader.fieldnames or [])
        id_col = column_map.get("id")
        if not id_col:
            return []
        return [_row_to_test_case(row, column_map) for row in reader if (row.get(id_col) or "").strip()]


def load_xlsx(path: str | Path, sheet: str | None = None) -> list[TestCase]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    column_map = _resolve_column_map(headers)
    id_col = column_map.get("id")
    if not id_col:
        return []

    cases = []
    for values in rows:
        row = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(values) if i < len(headers)}
        if (row.get(id_col) or "").strip():
            cases.append(_row_to_test_case(row, column_map))
    return cases


def load_test_cases(path: str | Path, sheet: str | None = None) -> list[TestCase]:
    """Dispatches to load_csv/load_xlsx based on file extension."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return load_csv(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return load_xlsx(path, sheet)
    raise ValueError(f"Unsupported file type: {path.suffix} (expected .csv, .xlsx, or .xlsm)")
